from flask import Flask, request, send_file
from openpyxl import load_workbook
import io

app = Flask(__name__)

@app.route('/unmerge', methods=['POST'])
def unmerge_and_download():
    if 'file' not in request.files:
        return 'No file uploaded.', 400

    file = request.files['file']
    in_mem_file = io.BytesIO(file.read())
    wb = load_workbook(in_mem_file)
    for ws in wb.worksheets:
        merged_ranges = list(ws.merged_cells.ranges)
        for cell_range in merged_ranges:
            ws.unmerge_cells(str(cell_range))
    out_mem_file = io.BytesIO()
    wb.save(out_mem_file)
    out_mem_file.seek(0)
    return send_file(
        out_mem_file,
        as_attachment=True,
        download_name='unmerged_' + file.filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == '__main__':
    app.run(debug=True)